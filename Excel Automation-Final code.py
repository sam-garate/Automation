#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd #https://automatetheboringstuff.com/chapter12/
import numpy as np
import os
import openpyxl as pyx


# In[2]:


book1 = pyx.load_workbook('Desktop\A.xlsx') #Parent workbook
book2 = pyx.load_workbook('Desktop\B.xlsx') #child workbook


# In[3]:


for i in book1:
    print(i.title)


# In[4]:


temp = book2['Sheet1']


# In[5]:


# Child banner: 
### column = 1 - table numbers (not needed in code)
### column = 2 - table & banner names
### column = 3 - Question number & text
### column = 4 - sub-question text 
### row that is in line with 'Table' = header 
### 1 row below the 'Table' row = sub-header

# Parent banner- 
### row 5 - Header
### row 6- sub-header
### column 1 - contains both questions & sub-questions


# In[6]:


#FRESH CODE - 6/3/2020  (F)

for rows1 in range(1,temp.max_row+1): #iterate over child file's rows
    table=temp.cell(row=rows1,column=2).value #lets check if there is a table starting in each row
    if table=='Table': #have to label the starting of each table in child banner as 'Table' to go identify that a table is starting
        #Function that gives the dimensions of the table in the child banners- 
        table_start=temp.cell(row=rows1,column=2)
        row_count=0 #rows 
        col_count=2 #columns- +1 to accomodate for the sig testing of last category, +1 to accomodate for the 1st column that we used to number tables
        #for rows: start counting at TABLE in column 2 till end of table. 
        for a in range(rows1, temp.max_row+1):
            if (temp.cell(row=a, column=2).value==None and temp.cell(row=a+1, column=2).value==None): #check if cell is not empty
                break
            else:
                row_count=row_count+1
        #for columns:start counting at Table in the top row till end of table+1
        for b in range(2,temp.max_column+1):
            if (temp.cell(row=rows1, column=b).value==None and temp.cell(row=rows1,column=b+1).value==None):
                break
            else:
                col_count = col_count+1
                
        for rows2 in range(rows1+2,rows1+row_count):
            banner=temp.cell(row=rows2,column=2).value
            question=temp.cell(row=rows2,column=3).value
            subq=temp.cell(row=rows2,column=4).value
            for i in book1.worksheets: #iterates over the multiple worksheets/banners in parent file
                if i.title==banner: #checks if the banner mentioned in child file exists in parent file
                    for rowsa in range(1,i.max_row+1): #iterates over rows in the parent banner opened
                        if question==i.cell(row=rowsa,column=1).value: #goes ahead only if question in child file is found in the parent banner
                            if i.cell(row=rowsa+4, column=1).value=='Total': #checks if cell has value 'Total' 4 cells below the cell containing the question in the parent banner
                                #Function that gives the dimensions of the table in the parent banners-
                                r_count=2 #count from cell containing 'total' till it encounters an empty cell, & add 2 to include the heading & subheading
                                c_count=1 #count from cell containing 'total' till last column, & add 1 to include sig testing in the last column
                                #for rows: match question in the parent banner, find the first total below it. the table starts 2 rows above the cell containing 'total' so as to contain the header & sub-header
                                for c in range(rowsa+4, i.max_row+1):
                                    if i.cell(row=c, column=1).value==None: #check if cell is not empty
                                        break
                                    else:
                                        r_count=r_count+1
                                                
                                #for columns: CORRECT
                                for d in range(1,i.max_column+1):
                                    if (i.cell(row=rowsa+4, column=d).value==None and i.cell(row=rowsa+4,column=d+1).value==None):
                                         break
                                    else:
                                        c_count = c_count+1
                                                
                            #Now comes the matching from child banner to parent banners- 
                            for rowsb in range(rowsa+4,rowsa+4+r_count):
                                if subq==i.cell(row=rowsb,column=1).value:
                                    for cols1 in range(5,col_count):
                                        header = temp.cell(row=table_start.row,column=cols1).value #2 rows above the row containing base numbers(total)
                                        sub_header = temp.cell(row=table_start.row+1,column=cols1).value #1 row above the row containing base numbers(total)
                                        for colsa in range(2,c_count):
                                            if header==i.cell(row=5,column=colsa).value:
                                                if sub_header==i.cell(row=6,column=colsa).value:
                                                    temp.cell(row=rows2,column=cols1).value=i.cell(row=rowsb,column=colsa).value #works till here
                                                    temp.cell(row=rows2,column=cols1+1).value=i.cell(row=rowsb,column=colsa+1).value


# In[7]:


book2.save('Desktop\C.xlsx')

